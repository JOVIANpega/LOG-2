# excel_writer.py
# 用途：將log分析結果匯出為Excel檔案，支援PASS/FAIL分頁，欄位依規格
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles.borders import Border, Side
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation
import re
import os

class ExcelWriter:
    def __init__(self):
        pass

    def _sanitize_cell_text(self, value: object) -> str:
        """清理欲寫入儲存格的文字：
        - 轉為字串
        - 移除非法控制字元 (openpyxl 限制)
        - 去除 ANSI/ESC 序列 (更廣泛)
        - 截斷過長文字 (3 萬字元)
        """
        try:
            text = '' if value is None else str(value)
            # 移除 Excel 禁用控制字元 (0x00-0x1F 無法接受，openpyxl 內建 regex)
            text = ILLEGAL_CHARACTERS_RE.sub('', text)
            # 移除常見 ANSI CSI 序列: ESC [ ... letter
            text = re.sub(r'\x1b\[[0-9;?]*[A-Za-z]', '', text)
            # 移除其他 ESC 開頭的短序列: ESC ... 單個字母終止
            text = re.sub(r'\x1b[^A-Za-z]{0,20}[A-Za-z]', '', text)
            # 移除顏色碼樣式殘餘 (保險)
            text = re.sub(r'\x1b', '', text)
            # 若首字為會被當作公式的危險字元，前綴 '\''
            if text and text[0] in ('=', '+', '-', '@'):
                text = "'" + text
            # 長度截斷 (32767為Excel上限，保留安全餘量)
            if len(text) > 30000:
                text = text[:30000]
            return text
        except Exception:
            return '' if value is None else str(value)[:30000]

    def export(self, pass_items, fail_items, output_path):
        """
        匯出分析結果到Excel，分為PASS/FAIL兩個sheet
        pass_items: List[dict]
        fail_items: List[dict]
        output_path: str
        """
        # PASS欄位：Step Name | 指令 | 回應 | 結果
        pass_df = pd.DataFrame(pass_items)
        if not pass_df.empty:
            cols = [c for c in ['file_name','step_name','command','response','result'] if c in pass_df.columns]
            pass_df = pass_df[cols]
            new_cols = ['檔名','Step Name','指令','回應','結果'] if 'file_name' in cols else ['Step Name','指令','回應','結果']
            pass_df.columns = new_cols
            # 清理
            for col in pass_df.columns:
                pass_df[col] = pass_df[col].apply(self._sanitize_cell_text)
        # FAIL欄位：Step Name | 指令 | 錯誤回應 | Retry 次數 | 錯誤原因
        fail_df = pd.DataFrame(fail_items)
        if not fail_df.empty:
            cols = [c for c in ['file_name','step_name','command','response','retry','error'] if c in fail_df.columns]
            fail_df = fail_df[cols]
            new_cols = ['檔名','Step Name','指令','錯誤回應','Retry 次數','錯誤原因'] if 'file_name' in cols else ['Step Name','指令','錯誤回應','Retry 次數','錯誤原因']
            fail_df.columns = new_cols
            # 清理
            for col in fail_df.columns:
                fail_df[col] = fail_df[col].apply(self._sanitize_cell_text)
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            if not pass_df.empty:
                pass_df.to_excel(writer, sheet_name='PASS', index=False)
            if not fail_df.empty:
                fail_df.to_excel(writer, sheet_name='FAIL', index=False)

    def export_with_summary(self, pass_items, fail_items, output_path):
        """
        產出 Summary/PASS/FAIL 三工作表；Summary 含每檔案一列的總覽與KPI。
        """
        pass_df = pd.DataFrame(pass_items)
        fail_df = pd.DataFrame(fail_items)
        # 正規化欄位
        if not pass_df.empty:
            if 'file_name' not in pass_df.columns:
                pass_df['file_name'] = ''
            pass_df = pass_df[['file_name','step_name','command','response','result']]
            pass_df.columns = ['檔名','Step Name','指令','回應','結果']
        if not fail_df.empty:
            if 'file_name' not in fail_df.columns:
                fail_df['file_name'] = ''
            fail_df = fail_df[['file_name','step_name','command','response','retry','error']]
            fail_df.columns = ['檔名','Step Name','指令','錯誤回應','Retry 次數','錯誤原因']
        # Summary 建立
        summary_rows = []
        all_files = set()
        if not pass_df.empty:
            all_files.update(pass_df['檔名'].unique())
        if not fail_df.empty:
            all_files.update(fail_df['檔名'].unique())
        for fn in sorted(all_files):
            p_cnt = 0 if pass_df.empty else (pass_df['檔名'] == fn).sum()
            f_cnt = 0 if fail_df.empty else (fail_df['檔名'] == fn).sum()
            result = 'FAIL' if f_cnt > 0 else 'PASS'
            # 簡要失敗原因（取第一個）
            fail_reason = ''
            if f_cnt > 0 and not fail_df.empty:
                subset = fail_df[fail_df['檔名'] == fn]
                fail_reason = subset['錯誤原因'].iloc[0] if not subset.empty else ''
            summary_rows.append({'檔名': fn, '結果': result, 'PASS筆數': p_cnt, 'FAIL筆數': f_cnt, '主要失敗原因': fail_reason})
        summary_df = pd.DataFrame(summary_rows)
        # KPI 概覽
        total_files = len(all_files)
        fail_files = (summary_df['結果'] == 'FAIL').sum() if not summary_df.empty else 0
        pass_files = total_files - fail_files
        kpi_df = pd.DataFrame([
            {'項目':'總檔案數','數值': total_files},
            {'項目':'PASS 檔數','數值': pass_files},
            {'項目':'FAIL 檔數','數值': fail_files},
            {'項目':'生成時間','數值': datetime.now().strftime('%Y-%m-%d %H:%M:%S')},
        ])
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            kpi_df.to_excel(writer, sheet_name='Summary', index=False, startrow=0)
            if not summary_df.empty:
                summary_df.to_excel(writer, sheet_name='Summary', index=False, startrow=len(kpi_df)+2)
            if not pass_df.empty:
                pass_df.to_excel(writer, sheet_name='PASS', index=False)
            if not fail_df.empty:
                fail_df.to_excel(writer, sheet_name='FAIL', index=False)

    # 依需求新增：生成 PASS匯總.xlsx 與 FAIL匯總.xlsx，含 Summary 與每檔案工作表
    def export_pass_fail_workbooks(self, folder_path: str, pass_logs: list, fail_logs: list):
        """
        輸出兩個活頁簿：
        - PASS匯總.xlsx：Summary + 每個 PASS LOG 的工作表
        - FAIL匯總.xlsx：Summary + 每個 FAIL LOG 的工作表
        pass_logs/fail_logs 需提供：
          [{
             'file_path': 絕對或相對路徑,
             'file_name': 檔名,
             'raw_lines': [原始行...],
             'ui_annotations': [{line_idx, line_content, color, ...}],
             'pass_items': [...],
             'fail_items': [...],
             'summary': { '測試日期時間': str, 'SFIS': 'ON'|'OFF', '測試總時間': str, 'FAIL原因': str可選 }
          }, ...]
        """
        pass_path = os.path.join(folder_path, 'PASS匯總.xlsx')
        fail_path = os.path.join(folder_path, 'FAIL匯總.xlsx')
        self._build_pass_workbook(pass_path, pass_logs)
        self._build_fail_workbook(fail_path, fail_logs)
        return pass_path, fail_path

    def _format_filename_with_timestamp(self, base_name: str) -> str:
        """將檔名中的連續14位時間戳 YYYYMMDDHHMMSS 轉為 YYYY-MMDD-HHMMSS 格式。找不到則原樣回傳。"""
        try:
            m = re.search(r'(20\d{12})', base_name)
            if not m:
                return base_name
            s = m.group(1)
            y, md, hms = s[:4], s[4:8], s[8:]
            return base_name.replace(s, f"{y}-{md}-{hms}")
        except Exception:
            return base_name

    def _build_preview_comment(self, entry: dict) -> str:
        """產生懸停預覽內容：顯示對應工作表名稱與原始LOG前幾行（加上簡易標記）。"""
        try:
            sheet_or_name = entry.get('file_name') or 'LOG'
            raw = entry.get('raw_lines') or []
            preview_lines = []
            header = f"對應工作表: {sheet_or_name}\n******** 預覽 ********"
            preview_lines.append(header)
            for i, line in enumerate(raw[:15], 1):
                s = str(line)
                # 簡易高亮標記
                if 'Do @STEP' in s:
                    s = f"[STEP] {s}"
                if 'FAIL' in s.upper() or 'ERROR' in s.upper():
                    s = f"[FAIL] {s}"
                s = re.sub(r'^\s*>\s*', '▶ ', s)
                s = re.sub(r'^\s*<\s*', '◀ ', s)
                preview_lines.append(s)
            preview_lines.append('************************')
            return '\n'.join(preview_lines)
        except Exception:
            return '對應工作表預覽不可用'

    def _add_input_prompt(self, ws, cell, title: str, message: str):
        """在儲存格上加一個資料驗證提示（白底有框），當選取時顯示。"""
        try:
            # Excel 對於訊息有長度限制，做適當截斷
            msg = (message or '')
            if len(msg) > 250:
                msg = msg[:250] + '…'
            dv = DataValidation(type="custom", formula1="TRUE", allow_blank=True, showInputMessage=True)
            dv.promptTitle = title[:30] if title else ''
            dv.prompt = msg
            ws.add_data_validation(dv)
            dv.add(cell.coordinate)
        except Exception:
            pass

    # 內部：PASS 匯總活頁簿
    def _build_pass_workbook(self, output_path: str, logs: list):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Summary'
        # 設定標籤顏色（紅色）
        try:
            ws.sheet_properties.tabColor = 'FFFF0000'
        except Exception:
            pass
        header_font = Font(name='Microsoft JhengHei', size=16, bold=True, color='FFFFFFFF')
        normal_font = Font(name='Microsoft JhengHei', size=10)
        center = Alignment(horizontal='center', vertical='center')
        deep_green = PatternFill('solid', fgColor='FF1B5E20')
        # 先建立各 LOG 原始工作表，並記錄 sheet 名稱
        sheet_map = {}
        for entry in logs:
            sheet_name_base = entry.get('file_name', 'LOG')
            sheet_name = self._unique_sheet_name(wb, sheet_name_base)
            sheet_map[entry.get('file_name')] = sheet_name
            ws2 = wb.create_sheet(title=sheet_name)
            cell = ws2.cell(row=1, column=1, value=self._sanitize_cell_text(entry.get('file_name')))
            cell.font = Font(name='Microsoft JhengHei', size=11, bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='FF2ECC71')
            cell.number_format = '@'
            step_labels = [str(i) for i, _ in enumerate(entry.get('pass_items') or [], 1)]
            ws2.cell(row=2, column=1, value=self._sanitize_cell_text('，'.join(step_labels) if step_labels else ''))
            ws2.cell(row=2, column=1).number_format = '@'
            ws2.cell(row=2, column=1).font = Font(name='Microsoft JhengHei', size=11)
            self._write_raw_log_with_annotations(ws2, start_row=3, raw_lines=entry.get('raw_lines') or [], annotations=entry.get('ui_annotations') or [], font=Font(name='Microsoft JhengHei', size=10), step_marks=entry.get('step_marks'))
            self._auto_fit_columns(ws2)
        # 以 PASS 項目清單分組
        def _pass_names(entry):
            names = []
            for it in (entry.get('pass_items') or []):
                nm = self._sanitize_cell_text(it.get('step_name',''))
                if nm:
                    names.append(nm)
            return tuple(names) if names else tuple()
        groups = {}
        for entry in logs:
            groups.setdefault(_pass_names(entry), []).append(entry)
        # 計算最大欄數（每25項一欄）
        CHUNK = 25
        max_chunks = 1
        for names in groups.keys():
            n = len(names)
            chunks = (n + CHUNK - 1) // CHUNK if n > 0 else 1
            if chunks > max_chunks:
                max_chunks = chunks
        # 標題列：檔名 + PASS項目 + PASS項目2..N
        headers = ['檔名', 'PASS項目'] + [f'PASS項目{i+1}' for i in range(1, max_chunks)]
        ws.append(headers)
        for c in range(1, len(headers)+1):
            cell = ws.cell(row=1, column=c)
            cell.font = header_font
            cell.alignment = center
            cell.fill = deep_green
        ws.freeze_panes = 'A2'
        thin = Side(border_style='thin', color='FF888888')
        thick = Side(border_style='thick', color='FF000000')
        start_data_row = ws.max_row + 1
        # 輸出每組一列
        for names, entries in groups.items():
            # 檔名欄：多LOG逐行，白底；加備註清單；不加超連結（避免誤判）
            lines = []
            for e in entries:
                base = self._sanitize_cell_text(e.get('file_name') or '')
                base_fmt = self._format_filename_with_timestamp(base)
                sfis = (e.get('summary') or {}).get('SFIS', '')
                sfis = (sfis or '').upper()
                secs = self._extract_total_secs(e.get('raw_lines') or [])
                sec_txt = f"測試總時間:{secs:.1f} Sec." if secs is not None else ''
                suffix = f"_SFIS_{sfis}" if sfis else ''
                lines.append(f"{base_fmt}{suffix} {sec_txt}".strip())
            display_name = '\n'.join(lines)
            r = ws.max_row + 1
            name_cell = ws.cell(row=r, column=1, value=self._sanitize_cell_text(display_name))
            name_cell.number_format='@'
            name_cell.font = Font(name='Microsoft JhengHei', size=10, color='FF000000')
            name_cell.alignment = Alignment(wrap_text=True, horizontal='left', vertical='top', shrink_to_fit=True)
            # 備註預覽對應sheet名
            sheet_list = [sheet_map.get(e.get('file_name')) for e in entries]
            preview = '\n- '.join([s for s in sheet_list if s][:8])
            # Summary 檔名欄不顯示懸停預覽與提示（避免誤判），僅在底部快速連結提供預覽/跳轉
            # PASS 項目：水平切成多欄
            numbered_all = [f"{i+1}. {nm}" for i, nm in enumerate(names)] if names else ['']
            for idx in range(max_chunks):
                chunk = numbered_all[idx*CHUNK:(idx+1)*CHUNK]
                txt = '' if not chunk else ('\n'.join(chunk) if idx == 0 else f"(PASS項目{idx+1})\n" + '\n'.join(chunk))
                c = 2 + idx
                cell = ws.cell(row=r, column=c, value=self._sanitize_cell_text(txt))
                cell.number_format='@'
                cell.font = Font(name='Microsoft JhengHei', size=10)
                cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left', shrink_to_fit=True)
            # 邊框（整列）
            last_col = 1 + max_chunks
            for c in range(1, last_col+1):
                ws.cell(row=r, column=c).border = Border(left=thin, right=thin, top=thin, bottom=thin)
        # 外框粗線
        end_r = ws.max_row
        last_col = 1 + max_chunks
        if end_r >= start_data_row:
            for c in range(1, last_col+1):
                ws.cell(row=start_data_row, column=c).border = ws.cell(row=start_data_row, column=c).border.copy(top=thick)
                ws.cell(row=end_r, column=c).border = ws.cell(row=end_r, column=c).border.copy(bottom=thick)
            for r in range(start_data_row, end_r+1):
                ws.cell(row=r, column=1).border = ws.cell(row=r, column=1).border.copy(left=thick)
                ws.cell(row=r, column=last_col).border = ws.cell(row=r, column=last_col).border.copy(right=thick)
        # 表格底部：SHEET 快速連結（點擊跳轉）
        link_title_row = ws.max_row + 2
        ws.cell(row=link_title_row, column=1, value='工作表快速連結（點擊跳轉）').font = Font(name='Microsoft JhengHei', size=11, bold=True)
        ws.cell(row=link_title_row, column=1).alignment = Alignment(horizontal='left')
        cur = link_title_row + 1
        for entry in logs:
            base = self._sanitize_cell_text(entry.get('file_name') or '')
            base_fmt = self._format_filename_with_timestamp(base)
            sfis = (entry.get('summary') or {}).get('SFIS', '')
            sfis = (sfis or '').upper()
            secs = self._extract_total_secs(entry.get('raw_lines') or [])
            sec_txt = f"測試總時間:{secs:.1f} Sec." if secs is not None else ''
            suffix = f"_SFIS_{sfis}" if sfis else ''
            display_name = f"{base_fmt}{suffix} {sec_txt}".strip()
            c = ws.cell(row=cur, column=1, value=self._sanitize_cell_text(display_name))
            c.number_format='@'
            c.font = Font(name='Microsoft JhengHei', size=11, color='FF0000FF', underline='single')
            c.alignment = Alignment(horizontal='left')
            sheet = sheet_map.get(entry.get('file_name'))
            if sheet:
                c.hyperlink = f"#'{sheet}'!A1"
                try:
                    c.comment = Comment(self._build_preview_comment(entry), 'LOG Analyzer')
                    c.comment.width = 400
                    c.comment.height = 500
                except Exception:
                    pass
                # 白底提示（提示箭頭在左上方，靠近視窗；Excel控制箭頭顯示位置有限）
                self._add_input_prompt(ws, c, '對應工作表', sheet)
            cur += 1
        # 更緊湊的欄寬
        min_widths = {1: 30}
        for i in range(2, last_col+1):
            min_widths[i] = 22
        self._auto_fit_columns(ws, min_widths=min_widths)
        wb.save(output_path)

    def _build_fail_workbook(self, output_path: str, logs: list):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Summary'
        # 設定標籤顏色（紅色）
        try:
            ws.sheet_properties.tabColor = 'FFFF0000'
        except Exception:
            pass
        header_font = Font(name='Microsoft JhengHei', size=16, bold=True, color='FFFFFFFF')
        normal_font = Font(name='Microsoft JhengHei', size=10)
        center = Alignment(horizontal='center', vertical='center')
        deep_green = PatternFill('solid', fgColor='FF1B5E20')
        headers = ['檔名', 'FAIL原因']
        ws.append(headers)
        for c in range(1, len(headers)+1):
            cell = ws.cell(row=1, column=c)
            cell.font = header_font
            cell.alignment = center
            cell.fill = deep_green
        ws.freeze_panes = 'A2'
        # 先建立各 LOG 原始工作表
        sheet_map = {}
        for entry in logs:
            sheet_name_base = entry.get('file_name', 'LOG')
            sheet_name = self._unique_sheet_name(wb, sheet_name_base)
            sheet_map[entry.get('file_name')] = sheet_name
            ws2 = wb.create_sheet(title=sheet_name)
            cell = ws2.cell(row=1, column=1, value=self._sanitize_cell_text(entry.get('file_name')))
            cell.font = Font(name='Microsoft JhengHei', size=11, bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='FFE74C3C')
            cell.number_format = '@'
            main_reason = (entry.get('summary') or {}).get('FAIL原因','')
            ws2.cell(row=2, column=1, value=self._sanitize_cell_text(main_reason)).number_format='@'
            ws2.cell(row=2, column=1).font = Font(name='Microsoft JhengHei', size=11)
            pass_steps = [str(i) for i, _ in enumerate(entry.get('pass_items') or [], 1)]
            ws2.cell(row=3, column=1, value=self._sanitize_cell_text('，'.join(pass_steps) if pass_steps else ''))
            ws2.cell(row=3, column=1).number_format='@'
            ws2.cell(row=3, column=1).font = Font(name='Microsoft JhengHei', size=11)
            self._write_raw_log_with_annotations(ws2, start_row=4, raw_lines=entry.get('raw_lines') or [], annotations=entry.get('ui_annotations') or [], font=Font(name='Microsoft JhengHei', size=10), step_marks=entry.get('step_marks'))
            self._auto_fit_columns(ws2)
        # Summary：以 FAIL 原因分組，每組一行
        thin = Side(border_style='thin', color='FF888888')
        thick = Side(border_style='thick', color='FF000000')
        # 以原因分組
        def _reason(entry):
            fails = entry.get('fail_items') or []
            return self._sanitize_cell_text(fails[0].get('error','')) if fails else ''
        r_groups = {}
        for e in logs:
            r_groups.setdefault(_reason(e), []).append(e)
        start_data_row = ws.max_row + 1
        for reason, entries in r_groups.items():
            # 檔名欄：多LOG逐行
            lines = []
            for e in entries:
                base = self._sanitize_cell_text(e.get('file_name') or '')
                base_fmt = self._format_filename_with_timestamp(base)
                sfis = (e.get('summary') or {}).get('SFIS','')
                sfis = (sfis or '').upper()
                secs = self._extract_total_secs(e.get('raw_lines') or [])
                sec_txt = f"測試總時間:{secs:.1f} Sec." if secs is not None else ''
                suffix = f"_SFIS_{sfis}" if sfis else ''
                lines.append(f"{base_fmt}{suffix} {sec_txt}".strip())
            display_name = '\n'.join(lines)
            r = ws.max_row + 1
            cell_name = ws.cell(row=r, column=1, value=self._sanitize_cell_text(display_name))
            cell_name.number_format='@'
            cell_name.font = Font(name='Microsoft JhengHei', size=10, color='FF000000')
            cell_name.alignment = Alignment(wrap_text=True, horizontal='left', vertical='top', shrink_to_fit=True)
            # 備註與超連結
            sheet_list = [sheet_map.get(e.get('file_name')) for e in entries]
            try:
                cell_name.comment = Comment(self._build_preview_comment(entries[0] if entries else {}), "LOG Analyzer")
                cell_name.comment.width = 400
                cell_name.comment.height = 500
            except Exception:
                pass
            if sheet_list and sheet_list[0]:
                cell_name.hyperlink = f"#'{sheet_list[0]}'!A1"
                # 白底提示
                self._add_input_prompt(ws, cell_name, '對應工作表', entries[0].get('file_name') or '')
            cell_reason = ws.cell(row=r, column=2, value=self._sanitize_cell_text(reason))
            cell_reason.number_format='@'
            cell_reason.font = Font(name='Microsoft JhengHei', size=10)
            cell_reason.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left', shrink_to_fit=True)
            for c in range(1, 2+1):
                ws.cell(row=r, column=c).border = Border(left=thin, right=thin, top=thin, bottom=thin)
        end_r = ws.max_row
        if end_r >= start_data_row:
            for c in range(1, 2+1):
                ws.cell(row=start_data_row, column=c).border = ws.cell(row=start_data_row, column=c).border.copy(top=thick)
                ws.cell(row=end_r, column=c).border = ws.cell(row=end_r, column=c).border.copy(bottom=thick)
            for r in range(start_data_row, end_r+1):
                ws.cell(row=r, column=1).border = ws.cell(row=r, column=1).border.copy(left=thick)
                ws.cell(row=r, column=2).border = ws.cell(row=r, column=2).border.copy(right=thick)
        # 表格底部：SHEET 快速連結
        link_title_row = ws.max_row + 2
        ws.cell(row=link_title_row, column=1, value='工作表快速連結（點擊跳轉）').font = Font(name='Microsoft JhengHei', size=10, bold=True)
        ws.cell(row=link_title_row, column=1).alignment = Alignment(horizontal='left')
        cur = link_title_row + 1
        for entry in logs:
            base = self._sanitize_cell_text(entry.get('file_name') or '')
            base_fmt = self._format_filename_with_timestamp(base)
            sfis = (entry.get('summary') or {}).get('SFIS','')
            sfis = (sfis or '').upper()
            secs = self._extract_total_secs(entry.get('raw_lines') or [])
            sec_txt = f"測試總時間:{secs:.1f} Sec." if secs is not None else ''
            suffix = f"_SFIS_{sfis}" if sfis else ''
            display_name = f"{base_fmt}{suffix} {sec_txt}".strip()
            c = ws.cell(row=cur, column=1, value=self._sanitize_cell_text(display_name))
            c.number_format='@'
            c.font = Font(name='Microsoft JhengHei', size=10, color='FF0000FF', underline='single')
            c.alignment = Alignment(horizontal='left')
            sheet = sheet_map.get(entry.get('file_name'))
            if sheet:
                c.hyperlink = f"#'{sheet}'!A1"
                try:
                    c.comment = Comment(self._build_preview_comment(entry), 'LOG Analyzer')
                    c.comment.width = 400
                    c.comment.height = 500
                except Exception:
                    pass
                self._add_input_prompt(ws, c, '對應工作表', sheet)
            cur += 1
        self._auto_fit_columns(ws, min_widths={1: 30, 2: 26})
        wb.save(output_path)

    def _write_raw_log_with_annotations(self, ws, start_row: int, raw_lines: list, annotations: list, font: Font, step_marks: dict | None = None):
        color_map = {
            'black': 'FF000000',
            'red': 'FFE74C3C',
            'green': 'FF2ECC71',
            'blue': 'FF3498DB',
            'purple': 'FF9B59B6',
        }
        marks = step_marks or {}
        # 以 annotations 的 color 欄位對應文字顏色；在步驟起始行前加上 1. 2. ... 標號
        for i, raw in enumerate(raw_lines, start=start_row):
            src_idx = i - start_row  # 對應原始 raw_lines 索引
            # 移除可能導致 openpyxl 錯誤的控制碼（例如 \x1b[0;41m 等ANSI碼或 0x00 字元）
            line = ILLEGAL_CHARACTERS_RE.sub('', str(raw))
            line = re.sub(r'\x1b\[[0-9;?]*[A-Za-z]', '', line)
            line = re.sub(r'\x1b[^A-Za-z]{0,20}[A-Za-z]', '', line)
            line = re.sub(r'\x1b', '', line)
            # 文字過長時截斷，避免 Excel 修復（32767 上限，預留安全餘量）
            if len(line) > 30000:
                line = line[:30000]
            mark = marks.get(src_idx)
            display = f"{mark}. {line}" if mark else line
            ws.cell(row=i, column=1, value=self._sanitize_cell_text(display))
            ws.cell(row=i, column=1).font = font
            ws.cell(row=i, column=1).number_format = '@'
            # 找到對應 annotation
            try:
                ann = annotations[src_idx] if src_idx < len(annotations) else None
            except Exception:
                ann = None
            if ann and ann.get('color'):
                hex_color = color_map.get(ann['color'], 'FF000000')
                ws.cell(row=i, column=1).font = Font(name=font.name, size=font.size, color=hex_color)

    def _auto_fit_columns(self, ws, min_widths: dict | None = None):
        max_width = {}
        for row in ws.iter_rows(values_only=True):
            for idx, cell in enumerate(row, 1):
                if cell is None:
                    continue
                # 增加可視 padding
                width = len(str(cell)) + 4
                if idx not in max_width or width > max_width[idx]:
                    max_width[idx] = width
        # 套用欄寬與下限
        for idx, width in max_width.items():
            base = min(120, width)
            if min_widths and idx in min_widths:
                base = max(base, min_widths[idx])
            ws.column_dimensions[get_column_letter(idx)].width = base

    def _safe_sheet_name(self, name: str) -> str:
        # Excel sheet 名稱限制處理
        invalid = '\\/*?:[]'
        safe = ''.join(['_' if ch in invalid else ch for ch in name])
        safe = safe.strip() or 'LOG'
        return safe[:31] if len(safe) > 31 else safe

    def _unique_sheet_name(self, wb, name: str) -> str:
        base = self._safe_sheet_name(name)
        if base not in wb.sheetnames:
            return base
        # 附加遞增後綴，保留31字元限制
        for i in range(2, 2000):
            suffix = f" ({i})"
            max_base_len = 31 - len(suffix)
            candidate = (base[:max_base_len] + suffix) if len(base) > max_base_len else (base + suffix)
            if candidate not in wb.sheetnames:
                return candidate
        # 極端情況，退回隨機尾碼
        import random
        return self._safe_sheet_name(f"{base}_{random.randint(1000,9999)}") 

    def _extract_total_secs(self, raw_lines: list):
        try:
            for i in range(len(raw_lines)-1, -1, -1):
                m = re.search(r'All phase Total Test Time\s*!\s*-+\s*([0-9]+(?:\.[0-9]+)?)\s*Sec', raw_lines[i], re.IGNORECASE)
                if m:
                    return float(m.group(1))
        except Exception:
            return None
        return None 